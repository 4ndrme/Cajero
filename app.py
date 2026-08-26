import os
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from config import Config, get_db_connection
from fpdf import FPDF
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import smtplib
import threading
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

DB_ENCRYPTION_KEY = os.getenv("DB_ENCRYPTION_KEY", "ClavePorDefecto123")

app = Flask(__name__)
app.config.from_object(Config)
app.secret_key = 'tu_clave_secreta_super_segura'  # Indispensable para las sesiones

# Vista del formulario de registro
@app.route('/registro')
def registro():
    return render_template('registro.html')

# API Endpoint para Registro de Cliente y Tarjeta
@app.route('/api/registrar_tarjeta', methods=['POST'])
def registrar_tarjeta():
    data = request.get_json()
    
    nombre = str(data.get('nombre', '')).strip()
    cedula = str(data.get('cedula', '')).strip()
    correo = str(data.get('correo', '')).strip()
    numero_cuenta = str(data.get('numero_cuenta', '')).strip()
    numero_tarjeta = str(data.get('numero_tarjeta', '')).strip()
    pin = str(data.get('pin', '')).strip()

    if not all([nombre, cedula, correo, numero_cuenta, numero_tarjeta, pin]):
        return jsonify({'success': False, 'message': 'Todos los campos (incluyendo el correo) son obligatorios.'}), 400

    if '@' not in correo or '.' not in correo:
        return jsonify({'success': False, 'message': 'Ingrese un correo electrónico válido.'}), 400

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("SELECT id FROM clientes WHERE cedula = %s", (cedula,))
        if cur.fetchone():
            return jsonify({'success': False, 'message': 'Ya existe un cliente con esta cédula.'}), 400

        cur.execute("SELECT id FROM tarjetas WHERE numero_tarjeta = %s", (numero_tarjeta,))
        if cur.fetchone():
            return jsonify({'success': False, 'message': 'Esta tarjeta ya está registrada.'}), 400

        cur.execute("""
            INSERT INTO clientes (nombre, cedula, correo) 
            VALUES (%s, %s, PGP_SYM_ENCRYPT(%s, %s)) RETURNING id
        """, (nombre, cedula, correo, DB_ENCRYPTION_KEY))
        cliente_id = cur.fetchone()[0]

        cur.execute("INSERT INTO cuentas (cliente_id, numero_cuenta, saldo) VALUES (%s, %s, 0.00) RETURNING id", (cliente_id, numero_cuenta))
        cuenta_id = cur.fetchone()[0]

        pin_hash = generate_password_hash(pin)
        cur.execute("INSERT INTO tarjetas (cuenta_id, numero_tarjeta, pin_hash) VALUES (%s, %s, %s)", (cuenta_id, numero_tarjeta, pin_hash))

        conn.commit()
        return jsonify({'success': True, 'message': 'Registro completado exitosamente.'})

    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'message': f'Error en el registro: {str(e)}'}), 500
    finally:
        cur.close()
        conn.close()

# 1. Pantalla de Inicio (Ingresar Tarjeta)
@app.route('/')
def index():
    session.clear()
    return render_template('index.html')

# Vista de la pantalla para ingresar el PIN
@app.route('/pin')
def vista_pin():
    if 'tarjeta_actual' not in session:
        return redirect(url_for('index'))
    return render_template('pin.html')
# 2. API Endpoint para validar si la tarjeta existe (CON BYPASS ADMIN)
@app.route('/api/validar_tarjeta', methods=['POST'])
def validar_tarjeta():
    data = request.get_json()
    numero_tarjeta = str(data.get('tarjeta', '')).strip()

    if not numero_tarjeta:
        return jsonify({'success': False, 'message': 'Ingrese un número de tarjeta.'}), 400

    # Lógica de Administrador
    if numero_tarjeta == '0000000000000000':
        session['tarjeta_actual'] = 'ADMIN'
        return jsonify({'success': True, 'redirect': '/pin'})

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("SELECT id, estado FROM tarjetas WHERE numero_tarjeta = %s", (numero_tarjeta,))
        tarjeta = cur.fetchone()

        if not tarjeta:
            return jsonify({'success': False, 'message': 'Tarjeta no encontrada. Regístrela primero.'}), 404

        if tarjeta[1] == 'Bloqueada':
            return jsonify({'success': False, 'message': 'La tarjeta se encuentra bloqueada.'}), 403

        session['tarjeta_actual'] = numero_tarjeta
        return jsonify({'success': True, 'redirect': '/pin'})

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
    finally:
        cur.close()
        conn.close()


# 4. API Endpoint Único para validar el PIN y manejar bloqueos (CON BYPASS ADMIN)
@app.route('/api/validar_pin', methods=['POST'])
def validar_pin():
    if 'tarjeta_actual' not in session:
        return jsonify({'success': False, 'message': 'Sesión expirada.'}), 401

    data = request.get_json()
    pin_ingresado = str(data.get('pin', '')).strip()
    numero_tarjeta = session['tarjeta_actual']

    # Lógica de Administrador
    if numero_tarjeta == 'ADMIN':
        if pin_ingresado == '9999':
            session['autenticado'] = True
            session['is_admin'] = True
            return jsonify({'success': True, 'redirect': '/admin_dashboard'})
        else:
            return jsonify({'success': False, 'message': 'PIN de administrador incorrecto.'}), 403

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("SELECT id, pin_hash, estado, intentos_fallidos FROM tarjetas WHERE numero_tarjeta = %s", (numero_tarjeta,))
        tarjeta_data = cur.fetchone()

        if not tarjeta_data:
            return jsonify({'success': False, 'message': 'Tarjeta no encontrada.'}), 404

        tarjeta_id, pin_hash, estado, intentos = tarjeta_data

        if estado == 'Bloqueada':
            return jsonify({'success': False, 'message': 'La tarjeta está bloqueada.'}), 403

        if check_password_hash(pin_hash, pin_ingresado):
            cur.execute("UPDATE tarjetas SET intentos_fallidos = 0 WHERE id = %s", (tarjeta_id,))
            conn.commit()
            
            session['autenticado'] = True
            return jsonify({'success': True, 'redirect': '/menu'})
        else:
            intentos += 1
            if intentos >= 3:
                cur.execute("UPDATE tarjetas SET intentos_fallidos = %s, estado = 'Bloqueada' WHERE id = %s", (intentos, tarjeta_id))
                conn.commit()
                session.clear()
                return jsonify({'success': False, 'message': 'PIN incorrecto. Su tarjeta ha sido bloqueada por seguridad.'}), 403
            else:
                cur.execute("UPDATE tarjetas SET intentos_fallidos = %s WHERE id = %s", (intentos, tarjeta_id))
                conn.commit()
                restantes = 3 - intentos
                return jsonify({'success': False, 'message': f'PIN incorrecto. Te quedan {restantes} intentos.'}), 400

    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'message': f'Error en servidor: {str(e)}'}), 500
    finally:
        cur.close()
        conn.close()

# 5. Pantalla del Menú Principal
@app.route('/menu')
def menu():
    if 'tarjeta_actual' not in session:
        return redirect(url_for('index'))
    return render_template('menu.html')

# 6. Pantalla de Saldo y Movimientos
@app.route('/saldo_movimientos')
def saldo_movimientos():
    if 'tarjeta_actual' not in session:
        return redirect(url_for('index'))
    
    numero_tarjeta = session['tarjeta_actual']
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT c.id, c.saldo 
        FROM cuentas c
        JOIN tarjetas t ON c.id = t.cuenta_id
        WHERE t.numero_tarjeta = %s
    """, (numero_tarjeta,))
    
    resultado = cur.fetchone()
    if not resultado:
        cur.close()
        conn.close()
        return redirect(url_for('index'))
        
    cuenta_id, saldo_actual = resultado
    
    cur.execute("INSERT INTO transacciones (cuenta_id, tipo, monto) VALUES (%s, 'Consulta', 0.00)", (cuenta_id,))
    conn.commit()
    
    cur.execute("""
        SELECT tipo, monto, fecha 
        FROM transacciones 
        WHERE cuenta_id = %s 
        ORDER BY fecha DESC 
        LIMIT 5
    """, (cuenta_id,))
    movimientos = cur.fetchall()
    
    cur.close()
    conn.close()
    
    return render_template('saldo_movimientos.html', saldo=saldo_actual, movimientos=movimientos)

# 7. Generación de PDF con Movimientos
@app.route('/imprimir_movimientos')
def imprimir_movimientos():
    if 'tarjeta_actual' not in session:
        return redirect(url_for('index'))
    
    numero_tarjeta = session['tarjeta_actual']
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT cl.nombre, c.numero_cuenta, c.saldo
        FROM clientes cl
        JOIN cuentas c ON cl.id = c.cliente_id
        JOIN tarjetas t ON c.id = t.cuenta_id
        WHERE t.numero_tarjeta = %s
    """, (numero_tarjeta,))
    cliente_info = cur.fetchone()
    
    cur.execute("""
        SELECT tipo, monto, fecha 
        FROM transacciones 
        WHERE cuenta_id = (SELECT cuenta_id FROM tarjetas WHERE numero_tarjeta = %s)
        ORDER BY fecha DESC
    """, (numero_tarjeta,))
    movimientos = cur.fetchall()
    
    cur.close()
    conn.close()

    if not cliente_info:
        return redirect(url_for('index'))

    pdf = FPDF()
    pdf.add_page()
    
    pdf.set_font("Arial", 'B', 18)
    pdf.set_text_color(25, 79, 43)
    pdf.cell(200, 10, txt="CajeBank", ln=True, align='C')
    
    pdf.set_font("Arial", 'B', 12)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(200, 10, txt="Estado de Cuenta y Movimientos", ln=True, align='C')
    pdf.ln(5)
    
    pdf.set_font("Arial", size=11)
    pdf.cell(200, 8, txt=f"Titular: {cliente_info[0]}", ln=True)
    pdf.cell(200, 8, txt=f"Nro. Cuenta: {cliente_info[1]}", ln=True)
    pdf.cell(200, 8, txt=f"Saldo Disponible: ${cliente_info[2]:.2f}", ln=True)
    pdf.ln(10)
    
    pdf.set_font("Arial", 'B', 11)
    pdf.set_fill_color(27, 138, 71)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(60, 10, "Fecha", border=1, fill=True, align='C')
    pdf.cell(70, 10, "Tipo de Movimiento", border=1, fill=True, align='C')
    pdf.cell(60, 10, "Monto", border=1, fill=True, align='C')
    pdf.ln()
    
    pdf.set_font("Arial", size=11)
    pdf.set_text_color(0, 0, 0)
    for mov in movimientos:
        fecha_str = mov[2].strftime('%Y-%m-%d %H:%M')
        tipo = mov[0]
        monto = float(mov[1])
        
        if monto == 0:
            monto_str = "----"
        elif tipo == 'Retiro' or 'Pago' in tipo:
            monto_str = f"- ${monto:.2f}"
        else:
            monto_str = f"+ ${monto:.2f}"
        
        pdf.cell(60, 10, fecha_str, border=1, align='C')
        pdf.cell(70, 10, tipo, border=1, align='C')
        pdf.cell(60, 10, monto_str, border=1, align='C')
        pdf.ln()

    pdf_output = pdf.output(dest='S').encode('latin1')
    return send_file(io.BytesIO(pdf_output), mimetype='application/pdf', as_attachment=True, download_name='CajeBank_Movimientos.pdf')

# 8. Pantalla de Retiro
@app.route('/retirar')
def retirar():
    if 'tarjeta_actual' not in session:
        return redirect(url_for('index'))
    return render_template('retirar.html')

# 9. API Endpoint Transaccional de Retiro
@app.route('/api/procesar_retiro', methods=['POST'])
def procesar_retiro():
    if 'tarjeta_actual' not in session:
        return jsonify({'success': False, 'message': 'Sesión expirada.'}), 401
    
    data = request.get_json()
    monto_retiro = float(data.get('monto'))

    if monto_retiro <= 0 or monto_retiro % 5 != 0:
        return jsonify({'success': False, 'message': 'El monto a retirar debe ser múltiplo de $5.00.'}), 400

    numero_tarjeta = session['tarjeta_actual']
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # 1. Validar fondos físicos en la bóveda de la máquina
        cur.execute("SELECT id, efectivo_disponible FROM cajero_estado ORDER BY id LIMIT 1 FOR UPDATE")
        estado_cajero = cur.fetchone()
        
        if not estado_cajero or float(estado_cajero[1]) < monto_retiro:
            return jsonify({'success': False, 'message': 'El cajero no cuenta con fondos físicos suficientes en este momento.'}), 400
            
        cajero_id = estado_cajero[0]
        nuevo_efectivo_boveda = float(estado_cajero[1]) - monto_retiro

        # 2. Validar fondos en la cuenta bancaria del cliente
        # 2. Validar fondos en la cuenta bancaria del cliente (AHORA TRAEMOS CORREO Y NOMBRE)
        cur.execute("""
            SELECT c.id, c.saldo, 
            PGP_SYM_DECRYPT(cl.correo::bytea, %s) AS correo_real, 
            cl.nombre 
            FROM cuentas c
            JOIN tarjetas t ON c.id = t.cuenta_id
            JOIN clientes cl ON c.cliente_id = cl.id
            WHERE t.numero_tarjeta = %s
            FOR UPDATE
        """, (DB_ENCRYPTION_KEY, numero_tarjeta))
        
        resultado = cur.fetchone()
        if not resultado:
            return jsonify({'success': False, 'message': 'Error de cuenta.'}), 404
            
        cuenta_id = resultado[0]
        saldo_actual = float(resultado[1])
        correo_cliente = resultado[2]  # <--- DATO OBTENIDO
        nombre_cliente = resultado[3]  # <--- DATO OBTENIDO
        
        if saldo_actual < monto_retiro:
            return jsonify({'success': False, 'message': 'Fondos insuficientes para esta transacción.'}), 400
        
        nuevo_saldo = saldo_actual - monto_retiro
        
        # 3. Ejecutar actualización triple
        cur.execute("UPDATE cuentas SET saldo = %s WHERE id = %s", (nuevo_saldo, cuenta_id))
        cur.execute("UPDATE cajero_estado SET efectivo_disponible = %s WHERE id = %s", (nuevo_efectivo_boveda, cajero_id))
        cur.execute("INSERT INTO transacciones (cuenta_id, tipo, monto) VALUES (%s, 'Retiro', %s)", (cuenta_id, monto_retiro))
        
        conn.commit()
        
        # 4. DISPARAR CORREO ASÍNCRONO DE FORMA INVISIBLE PARA EL USUARIO
        enviar_notificacion_asincrona(correo_cliente, nombre_cliente, "Retiro", monto_retiro)
        
        return jsonify({'success': True, 'message': f'Retiro exitoso de ${monto_retiro:.2f}. Retire su efectivo.'})
        
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'message': 'Fallo del sistema. La transacción ha sido revertida.'}), 500
    finally:
        cur.close()
        conn.close()
# 10. Pantalla de Otro Valor
@app.route('/otro_valor')
def otro_valor():
    if 'tarjeta_actual' not in session:
        return redirect(url_for('index'))
    return render_template('otro_valor.html')

# 11. Pantalla de Depósito
@app.route('/depositar')
def depositar():
    if 'tarjeta_actual' not in session:
        return redirect(url_for('index'))
    return render_template('depositar.html')

# 12. API Endpoint Transaccional de Depósito
@app.route('/api/procesar_deposito', methods=['POST'])
def procesar_deposito():
    if 'tarjeta_actual' not in session:
        return jsonify({'success': False, 'message': 'Sesión expirada.'}), 401

    data = request.get_json()
    monto_deposito = float(data.get('monto'))
    numero_tarjeta = session['tarjeta_actual'] 

    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # 1. Seleccionar y desencriptar
        cur.execute("""
            SELECT c.id, c.saldo, 
                   PGP_SYM_DECRYPT(cl.correo::bytea, %s) AS correo_real, 
                   cl.nombre 
            FROM cuentas c
            JOIN tarjetas t ON c.id = t.cuenta_id
            JOIN clientes cl ON c.cliente_id = cl.id
            WHERE t.numero_tarjeta = %s
            FOR UPDATE
        """, (DB_ENCRYPTION_KEY, numero_tarjeta))
        
        resultado = cur.fetchone()
        
        if not resultado:
            return jsonify({'success': False, 'message': 'Error de cuenta.'}), 404
            
        cuenta_id, saldo_actual_db, correo_cliente, nombre_cliente = resultado
        
        # 2. Convertir el saldo a float para evitar el Error 500
        nuevo_saldo = float(saldo_actual_db) + monto_deposito
        
        cur.execute("UPDATE cuentas SET saldo = %s WHERE id = %s", (nuevo_saldo, cuenta_id))
        cur.execute("INSERT INTO transacciones (cuenta_id, tipo, monto) VALUES (%s, 'Deposito', %s)", (cuenta_id, monto_deposito))
        
        # 3. Bóveda Física
        cur.execute("SELECT id, efectivo_disponible FROM cajero_estado ORDER BY id LIMIT 1 FOR UPDATE")
        estado_cajero = cur.fetchone()
        if estado_cajero:
            cur.execute("UPDATE cajero_estado SET efectivo_disponible = %s WHERE id = %s", (float(estado_cajero[1]) + monto_deposito, estado_cajero[0]))
            
        conn.commit()
        enviar_notificacion_asincrona(correo_cliente, nombre_cliente, "Depósito de Efectivo", monto_deposito)
        return jsonify({'success': True, 'message': f'Depósito exitoso de ${monto_deposito:.2f}.'})
        
    except Exception as e:
        conn.rollback()
        # Con este print, si algo falla, la consola te gritará exactamente el motivo
        print(f"❌ Error interno en depósito: {e}") 
        return jsonify({'success': False, 'message': 'Error en transacción.'}), 500
        
    finally:
        cur.close()
        conn.close()
# 15. Pantalla de Pago de Servicios
@app.route('/pagar_servicios')
def pagar_servicios():
    if 'tarjeta_actual' not in session:
        return redirect(url_for('index'))
    return render_template('pagar_servicios.html')
# 16. API Endpoint Transaccional para Pago de Servicios
@app.route('/api/procesar_pago_servicio', methods=['POST'])
def procesar_pago_servicio():
    if 'tarjeta_actual' not in session: 
        return jsonify({'success': False, 'message': 'Sesión expirada.'}), 401
    
    data = request.get_json()
    servicio, monto = data.get('servicio'), float(data.get('monto'))
    numero_tarjeta = session['tarjeta_actual']
    
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
        SELECT c.id, c.saldo, 
               PGP_SYM_DECRYPT(cl.correo::bytea, %s) AS correo_real, 
               cl.nombre 
        FROM cuentas c
        JOIN tarjetas t ON c.id = t.cuenta_id
        JOIN clientes cl ON c.cliente_id = cl.id
        WHERE t.numero_tarjeta = %s
        FOR UPDATE
        """, (DB_ENCRYPTION_KEY, numero_tarjeta))
        
        res = cur.fetchone()
        if not res: 
            return jsonify({'success': False, 'message': 'Error de cuenta.'}), 404
        
        cuenta_id, saldo_db, correo, nombre = res
        
        # MAGIA APLICADA: Convertimos el Decimal a float
        saldo_float = float(saldo_db)
        
        if saldo_float < monto: 
            return jsonify({'success': False, 'message': 'Fondos insuficientes.'}), 400
            
        cur.execute("UPDATE cuentas SET saldo = %s WHERE id = %s", (saldo_float - monto, cuenta_id))
        cur.execute("INSERT INTO transacciones (cuenta_id, tipo, monto) VALUES (%s, %s, %s)", (cuenta_id, f'Pago {servicio}', monto))
        conn.commit()
        
        enviar_notificacion_asincrona(correo, nombre, f"Pago de Servicio ({servicio})", monto)
        return jsonify({'success': True, 'message': f'Pago de {servicio} exitoso.'})
    except Exception as e:
        conn.rollback()
        print(f"❌ Error interno en pago de servicio: {e}")
        return jsonify({'success': False, 'message': 'Fallo en la transacción.'}), 500
    finally:
        cur.close()
        conn.close()

# 17. Generación de Comprobante de Pago de Servicio (PDF)
@app.route('/imprimir_recibo_servicio/<servicio>/<monto>')
def imprimir_recibo_servicio(servicio, monto):
    if 'tarjeta_actual' not in session:
        return redirect(url_for('index'))

    numero_tarjeta = session['tarjeta_actual']
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT cl.nombre, c.numero_cuenta, c.saldo
        FROM clientes cl
        JOIN cuentas c ON cl.id = c.cliente_id
        JOIN tarjetas t ON c.id = t.cuenta_id
        WHERE t.numero_tarjeta = %s
    """, (numero_tarjeta,))
    cliente_info = cur.fetchone()
    cur.close()
    conn.close()

    if not cliente_info:
        return redirect(url_for('index'))

    pdf = FPDF()
    pdf.add_page()
    
    pdf.set_font("Arial", 'B', 18)
    pdf.set_text_color(25, 79, 43)
    pdf.cell(200, 10, txt="CajeBank", ln=True, align='C')
    
    pdf.set_font("Arial", 'B', 14)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(200, 10, txt="Comprobante de Pago de Servicio", ln=True, align='C')
    pdf.ln(10)
    
    fecha_actual = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 8, txt=f"Fecha y Hora: {fecha_actual}", ln=True)
    pdf.cell(200, 8, txt=f"Titular de la Cuenta: {cliente_info[0]}", ln=True)
    pdf.cell(200, 8, txt=f"Nro. de Cuenta: {cliente_info[1]}", ln=True)
    pdf.ln(5)
    
    pdf.set_font("Arial", 'B', 12)
    pdf.cell(200, 8, txt=f"Servicio Pagado: {servicio}", ln=True)
    pdf.set_text_color(200, 0, 0)
    pdf.cell(200, 8, txt=f"Monto Debitado: ${float(monto):.2f}", ln=True)
    
    pdf.set_text_color(0, 0, 0)
    pdf.ln(5)
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 8, txt=f"Saldo Disponible: ${cliente_info[2]:.2f}", ln=True)
    
    pdf.ln(20)
    pdf.set_font("Arial", 'I', 10)
    pdf.cell(200, 10, txt="Gracias por confiar en CajeBank. Tu cajero, tu aliado.", ln=True, align='C')

    pdf_output = pdf.output(dest='S').encode('latin1')
    return send_file(io.BytesIO(pdf_output), mimetype='application/pdf', as_attachment=True, download_name=f'Comprobante_{servicio}.pdf')

# 18. Pantalla de Pago de Tarjeta de Crédito
@app.route('/pago_tarjeta')
def pago_tarjeta():
    if 'tarjeta_actual' not in session:
        return redirect(url_for('index'))
    return render_template('pago_tarjeta.html')
# 19. API Endpoint Transaccional para Pago de Tarjeta
@app.route('/api/procesar_pago_tarjeta', methods=['POST'])
def procesar_pago_tarjeta():
    if 'tarjeta_actual' not in session: 
        return jsonify({'success': False, 'message': 'Sesión expirada.'}), 401
    
    monto = float(request.get_json().get('monto'))
    numero_tarjeta = session['tarjeta_actual']
    
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT c.id, c.saldo, 
                   PGP_SYM_DECRYPT(cl.correo::bytea, %s) AS correo_real, 
                   cl.nombre 
            FROM cuentas c 
            JOIN tarjetas t ON c.id = t.cuenta_id 
            JOIN clientes cl ON c.cliente_id = cl.id
            WHERE t.numero_tarjeta = %s FOR UPDATE
        """, (DB_ENCRYPTION_KEY, numero_tarjeta))
        
        res = cur.fetchone()
        if not res: 
            return jsonify({'success': False, 'message': 'Error de cuenta.'}), 404
        
        cuenta_id, saldo_db, correo, nombre = res
        
        # MAGIA APLICADA: Convertimos el Decimal a float
        saldo_float = float(saldo_db)
        
        if saldo_float < monto: 
            return jsonify({'success': False, 'message': 'Fondos insuficientes.'}), 400
            
        cur.execute("UPDATE cuentas SET saldo = %s WHERE id = %s", (saldo_float - monto, cuenta_id))
        cur.execute("INSERT INTO transacciones (cuenta_id, tipo, monto) VALUES (%s, 'Pago Tarjeta', %s)", (cuenta_id, monto))
        conn.commit()
        
        enviar_notificacion_asincrona(correo, nombre, "Pago de Tarjeta de Crédito", monto)
        return jsonify({'success': True, 'message': 'Pago de tarjeta procesado exitosamente.'})
    except Exception as e:
        conn.rollback()
        print(f"❌ Error interno en pago de tarjeta: {e}")
        return jsonify({'success': False, 'message': 'Fallo en la transacción.'}), 500
    finally:
        cur.close()
        conn.close()
# 20. Generación de Comprobante PDF (Pago de Tarjeta)
@app.route('/imprimir_recibo_tarjeta/<monto>')
def imprimir_recibo_tarjeta(monto):
    if 'tarjeta_actual' not in session:
        return redirect(url_for('index'))

    numero_tarjeta = session['tarjeta_actual']
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT cl.nombre, c.numero_cuenta, c.saldo
        FROM clientes cl
        JOIN cuentas c ON cl.id = c.cliente_id
        JOIN tarjetas t ON c.id = t.cuenta_id
        WHERE t.numero_tarjeta = %s
    """, (numero_tarjeta,))
    cliente_info = cur.fetchone()
    cur.close()
    conn.close()

    if not cliente_info:
        return redirect(url_for('index'))

    pdf = FPDF()
    pdf.add_page()
    
    pdf.set_font("Arial", 'B', 18)
    pdf.set_text_color(25, 79, 43)
    pdf.cell(200, 10, txt="CajeBank", ln=True, align='C')
    
    pdf.set_font("Arial", 'B', 14)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(200, 10, txt="Comprobante de Pago de Tarjeta de Crédito", ln=True, align='C')
    pdf.ln(10)
    
    fecha_actual = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 8, txt=f"Fecha y Hora: {fecha_actual}", ln=True)
    pdf.cell(200, 8, txt=f"Titular de la Cuenta: {cliente_info[0]}", ln=True)
    pdf.cell(200, 8, txt=f"Nro. de Cuenta: {cliente_info[1]}", ln=True)
    pdf.ln(5)
    
    pdf.set_font("Arial", 'B', 12)
    pdf.set_text_color(200, 0, 0)
    pdf.cell(200, 8, txt=f"Monto Pagado: ${float(monto):.2f}", ln=True)
    
    pdf.set_text_color(0, 0, 0)
    pdf.ln(5)
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 8, txt=f"Saldo Disponible: ${cliente_info[2]:.2f}", ln=True)
    
    pdf.ln(20)
    pdf.set_font("Arial", 'I', 10)
    pdf.cell(200, 10, txt="Gracias por confiar en CajeBank. Tu cajero, tu aliado.", ln=True, align='C')

    pdf_output = pdf.output(dest='S').encode('latin1')
    return send_file(io.BytesIO(pdf_output), mimetype='application/pdf', as_attachment=True, download_name='Comprobante_Pago_Tarjeta.pdf')

# 21. Pantalla de Cambio de Clave
@app.route('/cambiar_clave')
def cambiar_clave():
    if 'tarjeta_actual' not in session:
        return redirect(url_for('index'))
    return render_template('cambiar_clave.html')

# 22. API Endpoint Transaccional para Cambio de Clave
@app.route('/api/procesar_cambio_clave', methods=['POST'])
def procesar_cambio_clave():
    if 'tarjeta_actual' not in session:
        return jsonify({'success': False, 'message': 'Sesión expirada.'}), 401

    data = request.get_json()
    cedula_ingresada = str(data.get('cedula', '')).strip()
    nuevo_pin = str(data.get('nuevo_pin', '')).strip()

    if not cedula_ingresada or not nuevo_pin or len(nuevo_pin) != 4 or not nuevo_pin.isdigit():
        return jsonify({'success': False, 'message': 'Datos inválidos. El PIN debe ser numérico de 4 dígitos.'}), 400

    numero_tarjeta = session['tarjeta_actual']
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        # Consulta transaccional uniendo Tarjetas -> Cuentas -> Clientes para obtener el correo y la cédula real
        cur.execute("""
            SELECT t.id, c.id, cl.correo, cl.nombre, cl.cedula
            FROM tarjetas t
            JOIN cuentas c ON t.cuenta_id = c.id
            JOIN clientes cl ON c.cliente_id = cl.id
            WHERE t.numero_tarjeta = %s
        """, (numero_tarjeta,))

        resultado = cur.fetchone()
        if not resultado:
            return jsonify({'success': False, 'message': 'Tarjeta no encontrada.'}), 404

        tarjeta_id, cuenta_id, correo_cliente, nombre_cliente, cedula_real = resultado

        # Validar que la cédula ingresada coincida con la de la base de datos
        if cedula_ingresada != cedula_real:
            return jsonify({'success': False, 'message': 'La cédula ingresada no coincide con el titular de la cuenta.'}), 403

        # Encriptar el nuevo PIN y actualizar la base de datos
        nuevo_hash = generate_password_hash(nuevo_pin)
        cur.execute("UPDATE tarjetas SET pin_hash = %s WHERE id = %s", (nuevo_hash, tarjeta_id))
        cur.execute("INSERT INTO transacciones (cuenta_id, tipo, monto) VALUES (%s, 'Cambio Clave', 0)", (cuenta_id,))

        conn.commit()

        # DISPARAR CORREO ASÍNCRONO
        enviar_notificacion_asincrona(correo_cliente, nombre_cliente, "Cambio de PIN de Seguridad", 0)

        return jsonify({'success': True, 'message': 'Su PIN ha sido actualizado con éxito.'})

    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'message': 'Error interno del servidor.'}), 500
    finally:
        cur.close()
        conn.close()

# 23. Panel de Administración
@app.route('/admin_dashboard')
def admin_dashboard():
    if not session.get('is_admin'):
        return redirect(url_for('index'))
        
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Consultar estado de la bóveda
    cur.execute("SELECT efectivo_disponible FROM cajero_estado ORDER BY id LIMIT 1")
    boveda = cur.fetchone()
    efectivo_boveda = float(boveda[0]) if boveda else 0.00
    
    cur.execute("SELECT COUNT(*) FROM transacciones")
    total_trx = cur.fetchone()[0]
    
    cur.execute("SELECT SUM(monto) FROM transacciones WHERE tipo = 'Retiro'")
    total_retiros = cur.fetchone()[0] or 0
    
    cur.execute("SELECT SUM(monto) FROM transacciones WHERE tipo = 'Deposito'")
    total_depositos = cur.fetchone()[0] or 0
    
    cur.execute("""
        SELECT tr.fecha, cl.nombre, tr.tipo, tr.monto
        FROM transacciones tr
        JOIN cuentas c ON tr.cuenta_id = c.id
        JOIN clientes cl ON c.cliente_id = cl.id
        ORDER BY tr.fecha DESC LIMIT 10
    """)
    recientes = cur.fetchall()

    cur.execute("""
            SELECT 
                cl.nombre, 
                cl.cedula, 
                PGP_SYM_DECRYPT(cl.correo::bytea, %s) AS correo_real,
                c.numero_cuenta,
                t.numero_tarjeta,
                c.saldo,
                t.estado
            FROM clientes cl
            JOIN cuentas c ON cl.id = c.cliente_id
            JOIN tarjetas t ON c.id = t.cuenta_id
            ORDER BY cl.nombre ASC
        """, (DB_ENCRYPTION_KEY,))
        
    clientes_registrados = cur.fetchall()
    
    cur.close()
    conn.close()

    return render_template('admin_dashboard.html', 
                                total_trx=total_trx,
                                total_retiros=total_retiros,
                                total_depositos=total_depositos,
                                efectivo_boveda=efectivo_boveda,
                                recientes=recientes,
                                clientes=clientes_registrados)
# 24. Exportar Auditoría a Excel (Con Estilo Mejorado)
@app.route('/descargar_excel')
def descargar_excel():
    if not session.get('is_admin'):
        return redirect(url_for('index'))

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT tr.fecha, cl.nombre, cl.cedula, tr.tipo, tr.monto
        FROM transacciones tr
        JOIN cuentas c ON tr.cuenta_id = c.id
        JOIN clientes cl ON c.cliente_id = cl.id
        ORDER BY tr.fecha DESC
    """)
    movimientos = cur.fetchall()
    cur.close()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auditoria Cajero"
    
    # 1. Definir Estilos
    header_font = Font(bold=True, color="FFFFFF")
    # Color verde oscuro acorde al diseño de la app
    header_fill = PatternFill("solid", fgColor="0B5345") 
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'), 
        right=Side(style='thin', color='CCCCCC'), 
        top=Side(style='thin', color='CCCCCC'), 
        bottom=Side(style='thin', color='CCCCCC')
    )

    # 2. Insertar y Estilizar Cabeceras
    headers = ["Fecha y Hora", "Cliente", "Cédula", "Acción Realizada", "Monto ($)"]
    ws.append(headers)
    
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 3. Insertar Datos y Formatear Filas
    for row_idx, mov in enumerate(movimientos, start=2):
        fecha_str = mov[0].strftime('%Y-%m-%d %H:%M:%S')
        monto_float = float(mov[4])
        
        ws.append([fecha_str, mov[1], mov[2], mov[3], monto_float])
        
        # Aplicar alineación y bordes
        ws.cell(row=row_idx, column=1).alignment = center_align
        ws.cell(row=row_idx, column=2).alignment = left_align
        ws.cell(row=row_idx, column=3).alignment = center_align
        ws.cell(row=row_idx, column=4).alignment = center_align
        
        # Formato de Moneda para la columna de Monto
        monto_cell = ws.cell(row=row_idx, column=5)
        monto_cell.alignment = right_align
        if monto_float == 0:
            monto_cell.value = "----"
            monto_cell.alignment = center_align
        else:
            monto_cell.number_format = '"$"#,##0.00'

        for col_idx in range(1, 6):
            ws.cell(row=row_idx, column=col_idx).border = thin_border

    # 4. Ajustar Ancho de Columnas Fijas
    column_widths = {
        'A': 22, # Fecha y Hora
        'B': 30, # Cliente
        'C': 15, # Cédula
        'D': 25, # Acción
        'E': 15  # Monto
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Convertir libro a formato descargable en memoria
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    return send_file(
        out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='Auditoria_CajeBank.xlsx'
    )

@app.route('/api/abastecer_cajero', methods=['POST'])
def abastecer_cajero():
    if not session.get('is_admin'):
        return jsonify({'success': False, 'message': 'Acceso denegado.'}), 403

    data = request.get_json()
    monto_recarga = float(data.get('monto', 10000.00))

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, efectivo_disponible FROM cajero_estado ORDER BY id LIMIT 1 FOR UPDATE")
        estado = cur.fetchone()
        
        if estado:
            nuevo_monto = float(estado[1]) + monto_recarga
            cur.execute("UPDATE cajero_estado SET efectivo_disponible = %s WHERE id = %s", (nuevo_monto, estado[0]))
        else:
            cur.execute("INSERT INTO cajero_estado (efectivo_disponible) VALUES (%s)", (monto_recarga,))
            
        conn.commit()
        return jsonify({'success': True, 'message': f'La máquina ha sido abastecida con ${monto_recarga:,.2f} físicos.'})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'message': 'Fallo al abastecer la bóveda física.'}), 500
    finally:
        cur.close()
        conn.close()

# Función asíncrona para enviar notificaciones por correo
def enviar_notificacion_asincrona(correo_destino, cliente, tipo_operacion, monto):
    def enviar():
        try:
            smtp_server = os.getenv("SMTP_SERVER", "smtp.gmail.com")
            smtp_port = int(os.getenv("SMTP_PORT", 465))
            smtp_user = os.getenv("SMTP_USER")
            smtp_pass = os.getenv("SMTP_PASS")

            if not smtp_user or not smtp_pass:
                print("Credenciales SMTP no configuradas. Correo cancelado.")
                return

            msg = MIMEMultipart()
            msg['From'] = f"CajeBank Notificaciones <{smtp_user}>"
            msg['To'] = correo_destino
            msg['Subject'] = f"CajeBank - Confirmación de {tipo_operacion}"

            monto_str = "----" if monto == 0 else f"${monto:.2f}"

            cuerpo = f"""
            Hola {cliente},

            Te notificamos que se ha procesado exitosamente una nueva transacción en tu cuenta de CajeBank.

            DETALLES DE LA OPERACIÓN:
            ----------------------------------------
            - Acción Realizada: {tipo_operacion}
            - Monto Involucrado: {monto_str}
            ----------------------------------------

            Si no reconoces esta actividad, por favor acércate a una sucursal física de inmediato.
            
            Gracias por confiar en CajeBank.
            """
            
            msg.attach(MIMEText(cuerpo, 'plain'))

            server = smtplib.SMTP_SSL(smtp_server, smtp_port)
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)
            server.quit()
            print(f"[SMTP] Alerta enviada con éxito a {correo_destino}")
            
        except Exception as e:
            print(f"[SMTP] Error al intentar enviar el correo: {e}")

    # Disparar el hilo en segundo plano para no bloquear el frontend
    hilo = threading.Thread(target=enviar)
    hilo.start()

# Logout
@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True, port=5000)